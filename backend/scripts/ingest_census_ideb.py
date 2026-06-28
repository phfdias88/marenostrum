"""
IDEB municipal (INEP 2023) — anos iniciais e finais do Ensino Fundamental,
rede Pública. Download dos ZIPs do INEP (xlsx), parse com openpyxl.

Acha as colunas pelo cabeçalho técnico (CO_MUNICIPIO, REDE, VL_OBSERVADO_2023)
em vez de índice fixo — robusto a mudança de layout.

Rodar: docker compose exec -T api sh -c 'PYTHONPATH=/app python /tmp/ingest_ideb.py'
"""
import io
import ssl
import urllib.request
import zipfile

import openpyxl
from sqlalchemy import text

from app.core.database import SessionLocal

# O servidor do INEP (download.inep.gov.br) serve uma cadeia de certificado que
# o CA bundle do container não valida (intermediário ICP-Brasil faltando). Como
# o arquivo é PÚBLICO (estatística aberta, sem dado sensível), baixamos com
# verificação desligada — só para esta fonte.
_SSL_NOVERIFY = ssl.create_default_context()
_SSL_NOVERIFY.check_hostname = False
_SSL_NOVERIFY.verify_mode = ssl.CERT_NONE

URLS = {
    "ini": "https://download.inep.gov.br/ideb/resultados/divulgacao_anos_iniciais_municipios_2023.zip",
    "fin": "https://download.inep.gov.br/ideb/resultados/divulgacao_anos_finais_municipios_2023.zip",
}


_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")


def _download(url: str) -> bytes:
    """Baixa com UA de browser + retry (o INEP reseta conexões de UA não-browser)."""
    import time
    last = None
    for attempt in range(4):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": _UA, "Accept": "*/*", "Connection": "keep-alive",
            })
            with urllib.request.urlopen(req, timeout=300, context=_SSL_NOVERIFY) as r:
                return r.read()
        except Exception as e:  # noqa: BLE001
            last = e
            print(f"  tentativa {attempt+1} falhou: {e}; retry...")
            time.sleep(5 * (attempt + 1))
    raise last


def fetch_ideb(url: str) -> dict[str, float]:
    """cd_mun(7) -> IDEB observado 2023, só rede Pública."""
    blob = _download(url)
    zf = zipfile.ZipFile(io.BytesIO(blob))
    name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx"))
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    out: dict[str, float] = {}
    col = {}  # nome técnico -> índice
    for row in ws.iter_rows(values_only=True):
        if not col:
            # acha a linha de cabeçalho técnico (tem CO_MUNICIPIO)
            vals = [str(c).strip() if c is not None else "" for c in row]
            if "CO_MUNICIPIO" in vals:
                for i, v in enumerate(vals):
                    col[v] = i
            continue
        cod = row[col["CO_MUNICIPIO"]]
        rede = str(row[col.get("REDE", -1)] or "")
        if cod is None or not rede.startswith("P"):  # latin-1: 'Pública'~'P...'
            continue
        ideb = row[col.get("VL_OBSERVADO_2023", -1)]
        if isinstance(ideb, (int, float)):
            out[str(int(cod))] = float(ideb)
    return out


def main() -> None:
    res = {}
    for key, url in URLS.items():
        print(f"baixando IDEB {key}: {url}")
        res[key] = fetch_ideb(url)
        print(f"  {key}: {len(res[key])} municípios (rede pública)")

    db = SessionLocal()
    try:
        muns = [r[0] for r in db.execute(
            text("SELECT cd_mun FROM census_geo WHERE level='municipio'")
        ).all()]
        upd = text(
            "UPDATE census_geo SET ideb_anos_iniciais=:i, ideb_anos_finais=:f "
            "WHERE level='municipio' AND cd_mun=:cd"
        )
        n = 0
        for cd in muns:
            i, f = res["ini"].get(cd), res["fin"].get(cd)
            if i is None and f is None:
                continue
            db.execute(upd, {"cd": cd, "i": i, "f": f})
            n += 1
        db.commit()
        print(f"municípios atualizados com IDEB: {n}")
        sample = db.execute(text(
            "SELECT nm_mun, ideb_anos_iniciais, ideb_anos_finais FROM census_geo "
            "WHERE level='municipio' AND ideb_anos_iniciais IS NOT NULL "
            "ORDER BY ideb_anos_iniciais DESC LIMIT 5"
        )).all()
        print("top 5 IDEB anos iniciais:")
        for s in sample:
            print(f"  {s[0]}: iniciais {s[1]} | finais {s[2]}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
